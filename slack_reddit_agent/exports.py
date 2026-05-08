import json
import re
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .config import EXPORTS_DIR, MAX_EXPORT_ROWS


def sanitize_filename(value: str, fallback: str) -> str:
    """Create a filesystem-safe filename stem."""
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip()).strip("._")
    return cleaned or fallback


def sanitize_sheet_name(value: str) -> str:
    """Create an Excel-safe worksheet title."""
    cleaned = re.sub(r"[:\\\\/?*\\[\\]]", " ", value).strip()
    return (cleaned or "Report")[:31]


def coerce_excel_value(value: Any) -> Any:
    """Normalize nested data for a worksheet cell."""
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=True)


def _normalize_export_payload(payload: dict[str, Any]) -> dict[str, Any] | None:
    export_type = str(payload.get("export_type") or "").strip().lower()

    if export_type == "docx":
        paragraphs = payload.get("paragraphs")
        if not isinstance(paragraphs, list):
            return None
        return payload

    rows = payload.get("rows")
    columns = payload.get("columns")
    if isinstance(rows, list) and isinstance(columns, list):
        normalized = dict(payload)
        normalized.setdefault("export_type", "xlsx")
        return normalized

    return None


def extract_export_payload(text: str) -> tuple[str, dict[str, Any] | None]:
    """Extract a JSON export block from an agent response if present."""
    code_block_pattern = re.compile(r"```json\s*(\{.*?\})\s*```", re.DOTALL)

    for match in code_block_pattern.finditer(text):
        candidate = match.group(1).strip()
        try:
            payload = json.loads(candidate)
        except json.JSONDecodeError:
            continue

        if not isinstance(payload, dict):
            continue

        normalized_payload = _normalize_export_payload(payload)
        if normalized_payload is None:
            continue

        message_text = text.replace(match.group(0), "").strip()
        return message_text, normalized_payload

    return text, None


def create_excel_report(payload: dict[str, Any]) -> Path:
    """Build an XLSX file from the agent export payload."""
    workbook_name = sanitize_filename(
        str(payload.get("workbook_name") or "reddit_report"),
        fallback="reddit_report",
    )
    sheet_name = sanitize_sheet_name(str(payload.get("sheet_name") or "Report"))
    summary = str(payload.get("summary") or "").strip()
    columns = [str(column).strip() for column in payload.get("columns", []) if str(column).strip()]
    raw_rows = payload.get("rows") or []
    row_limit = payload.get("row_limit", MAX_EXPORT_ROWS)

    rows: list[dict[str, Any]] = []
    row_source = raw_rows if row_limit is None else raw_rows[: int(row_limit)]
    for raw_row in row_source:
        if isinstance(raw_row, dict):
            rows.append(raw_row)

    if not columns and rows:
        columns = list(rows[0].keys())

    if not columns:
        columns = ["result"]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(columns)

    for header_cell in worksheet[1]:
        header_cell.font = Font(bold=True)

    for row in rows:
        worksheet.append([coerce_excel_value(row.get(column, "")) for column in columns])

    for index, column in enumerate(columns, start=1):
        cell_values = [str(column)]
        for row in rows:
            cell_values.append(str(coerce_excel_value(row.get(column, ""))))
        width = min(max(len(value) for value in cell_values) + 2, 60)
        worksheet.column_dimensions[get_column_letter(index)].width = width

    if summary:
        summary_sheet = workbook.create_sheet(title="Summary")
        summary_sheet["A1"] = "Summary"
        summary_sheet["A1"].font = Font(bold=True)
        summary_sheet["A2"] = summary
        summary_sheet.column_dimensions["A"].width = 100

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = EXPORTS_DIR / f"{workbook_name}_{timestamp}.xlsx"
    workbook.save(file_path)
    return file_path


def _coerce_docx_paragraphs(payload: dict[str, Any]) -> list[str]:
    paragraphs: list[str] = []
    summary = str(payload.get("summary") or "").strip()
    if summary:
        paragraphs.append(summary)

    for raw_paragraph in payload.get("paragraphs", []):
        text = str(raw_paragraph or "").strip()
        if text:
            paragraphs.append(text)

    if not paragraphs:
        paragraphs.append("No content available.")
    return paragraphs


def _build_docx_document_xml(title: str, paragraphs: list[str]) -> str:
    body_parts: list[str] = []
    if title:
        body_parts.append(
            "<w:p><w:r><w:rPr><w:b/></w:rPr><w:t>"
            f"{escape(title)}"
            "</w:t></w:r></w:p>"
        )

    for paragraph in paragraphs:
        lines = str(paragraph).splitlines() or [""]
        paragraph_xml = ["<w:p>"]
        for line_index, line in enumerate(lines):
            if line_index:
                paragraph_xml.append("<w:r><w:br/></w:r>")
            text_value = line if line else " "
            paragraph_xml.append(
                "<w:r><w:t xml:space=\"preserve\">"
                f"{escape(text_value)}"
                "</w:t></w:r>"
            )
        paragraph_xml.append("</w:p>")
        body_parts.append("".join(paragraph_xml))

    body_xml = "".join(body_parts)
    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<w:document xmlns:wpc=\"http://schemas.microsoft.com/office/word/2010/wordprocessingCanvas\" "
        "xmlns:mc=\"http://schemas.openxmlformats.org/markup-compatibility/2006\" "
        "xmlns:o=\"urn:schemas-microsoft-com:office:office\" "
        "xmlns:r=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships\" "
        "xmlns:m=\"http://schemas.openxmlformats.org/officeDocument/2006/math\" "
        "xmlns:v=\"urn:schemas-microsoft-com:vml\" "
        "xmlns:wp14=\"http://schemas.microsoft.com/office/word/2010/wordprocessingDrawing\" "
        "xmlns:wp=\"http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing\" "
        "xmlns:w10=\"urn:schemas-microsoft-com:office:word\" "
        "xmlns:w=\"http://schemas.openxmlformats.org/wordprocessingml/2006/main\" "
        "xmlns:w14=\"http://schemas.microsoft.com/office/word/2010/wordml\" "
        "xmlns:w15=\"http://schemas.microsoft.com/office/word/2012/wordml\" "
        "xmlns:wpg=\"http://schemas.microsoft.com/office/word/2010/wordprocessingGroup\" "
        "xmlns:wpi=\"http://schemas.microsoft.com/office/word/2010/wordprocessingInk\" "
        "xmlns:wne=\"http://schemas.microsoft.com/office/word/2006/wordml\" "
        "xmlns:wps=\"http://schemas.microsoft.com/office/word/2010/wordprocessingShape\" "
        "mc:Ignorable=\"w14 w15 wp14\">"
        "<w:body>"
        f"{body_xml}"
        "<w:sectPr><w:pgSz w:w=\"12240\" w:h=\"15840\"/><w:pgMar w:top=\"1440\" w:right=\"1440\" "
        "w:bottom=\"1440\" w:left=\"1440\" w:header=\"708\" w:footer=\"708\" w:gutter=\"0\"/>"
        "<w:cols w:space=\"708\"/><w:docGrid w:linePitch=\"360\"/></w:sectPr>"
        "</w:body></w:document>"
    )


def create_docx_report(payload: dict[str, Any]) -> Path:
    """Build a DOCX file from a document export payload."""
    document_name = sanitize_filename(
        str(payload.get("document_name") or payload.get("workbook_name") or "document_export"),
        fallback="document_export",
    )
    document_title = str(payload.get("title") or document_name).strip()
    paragraphs = _coerce_docx_paragraphs(payload)

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = EXPORTS_DIR / f"{document_name}_{timestamp}.docx"

    content_types_xml = (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<Types xmlns=\"http://schemas.openxmlformats.org/package/2006/content-types\">"
        "<Default Extension=\"rels\" ContentType=\"application/vnd.openxmlformats-package.relationships+xml\"/>"
        "<Default Extension=\"xml\" ContentType=\"application/xml\"/>"
        "<Override PartName=\"/word/document.xml\" "
        "ContentType=\"application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml\"/>"
        "<Override PartName=\"/docProps/core.xml\" "
        "ContentType=\"application/vnd.openxmlformats-package.core-properties+xml\"/>"
        "<Override PartName=\"/docProps/app.xml\" "
        "ContentType=\"application/vnd.openxmlformats-officedocument.extended-properties+xml\"/>"
        "</Types>"
    )
    rels_xml = (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<Relationships xmlns=\"http://schemas.openxmlformats.org/package/2006/relationships\">"
        "<Relationship Id=\"rId1\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument\" "
        "Target=\"word/document.xml\"/>"
        "<Relationship Id=\"rId2\" Type=\"http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties\" "
        "Target=\"docProps/core.xml\"/>"
        "<Relationship Id=\"rId3\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties\" "
        "Target=\"docProps/app.xml\"/>"
        "</Relationships>"
    )
    document_rels_xml = (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<Relationships xmlns=\"http://schemas.openxmlformats.org/package/2006/relationships\"/>"
    )
    core_xml = (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<cp:coreProperties xmlns:cp=\"http://schemas.openxmlformats.org/package/2006/metadata/core-properties\" "
        "xmlns:dc=\"http://purl.org/dc/elements/1.1/\" "
        "xmlns:dcterms=\"http://purl.org/dc/terms/\" "
        "xmlns:dcmitype=\"http://purl.org/dc/dcmitype/\" "
        "xmlns:xsi=\"http://www.w3.org/2001/XMLSchema-instance\">"
        f"<dc:title>{escape(document_title)}</dc:title>"
        "<dc:creator>Reddit Agent</dc:creator>"
        "</cp:coreProperties>"
    )
    app_xml = (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<Properties xmlns=\"http://schemas.openxmlformats.org/officeDocument/2006/extended-properties\" "
        "xmlns:vt=\"http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes\">"
        "<Application>Reddit Agent</Application>"
        "</Properties>"
    )
    document_xml = _build_docx_document_xml(document_title, paragraphs)

    with zipfile.ZipFile(file_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types_xml)
        archive.writestr("_rels/.rels", rels_xml)
        archive.writestr("word/document.xml", document_xml)
        archive.writestr("word/_rels/document.xml.rels", document_rels_xml)
        archive.writestr("docProps/core.xml", core_xml)
        archive.writestr("docProps/app.xml", app_xml)

    return file_path


def create_export_file(payload: dict[str, Any]) -> Path:
    """Build the appropriate export file for a payload."""
    export_type = str(payload.get("export_type") or "xlsx").strip().lower()
    if export_type == "docx":
        return create_docx_report(payload)
    return create_excel_report(payload)


def export_title(payload: dict[str, Any]) -> str:
    """Return the best human-readable title for the export."""
    export_type = str(payload.get("export_type") or "xlsx").strip().lower()
    if export_type == "docx":
        return str(payload.get("title") or payload.get("document_name") or "document_export")
    return str(payload.get("workbook_name") or "reddit_report")


def export_success_message(payload: dict[str, Any]) -> str:
    """Return a compact success message for the created export."""
    export_type = str(payload.get("export_type") or "xlsx").strip().lower()
    if export_type == "docx":
        return "Document file created and uploaded."
    return "Excel report created and uploaded."
