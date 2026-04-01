import asyncio
import atexit
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from agents import Agent, Runner, SQLiteSession
from composio import Composio
from composio_openai_agents import OpenAIAgentsProvider
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from slack_bolt.adapter.socket_mode.async_handler import AsyncSocketModeHandler
from slack_bolt.async_app import AsyncApp

# Load environment variables from the project .env file.
load_dotenv()

# Initialize Composio for OpenAI Agents.
composio = Composio(provider=OpenAIAgentsProvider())

user_sessions: dict[str, Any] = {}
memory_sessions: dict[str, SQLiteSession] = {}

LOCK_FILE = ".slack_bot.lock"
DEFAULT_MODEL = os.getenv("OPENAI_MODEL", "gpt-5.2")
DEFAULT_TOOLKITS = [
    toolkit.strip()
    for toolkit in os.getenv("COMPOSIO_TOOLKITS", "reddit").split(",")
    if toolkit.strip()
]
SHARED_COMPOSIO_USER_ID = os.getenv("COMPOSIO_SHARED_USER_ID", "").strip()
SHARED_CONNECTED_ACCOUNT_ID = os.getenv(
    "COMPOSIO_SHARED_CONNECTED_ACCOUNT_ID", ""
).strip()
ENV_ALLOWED_USERS = {
    user_id.strip()
    for user_id in os.getenv("SLACK_ALLOWED_USERS", "").split(",")
    if user_id.strip()
}
SLACK_MESSAGE_LIMIT = 3500
SLACK_BLOCK_TEXT_LIMIT = 2900
SLACK_HEADER_TEXT_LIMIT = 150
SLACK_CONTEXT_MESSAGE_LIMIT = 8
SLACK_CONTEXT_CHAR_LIMIT = 3000
MAX_EXPORT_ROWS = 200
EXPORTS_DIR = Path("exports")
ACCESS_CONTROL_FILE = Path(".slack_access_control.json")
REDDIT_BRAND_NAME = os.getenv("REDDIT_BRAND_NAME", "SkinPal").strip() or "SkinPal"
REDDIT_BRAND_CONTEXT = os.getenv("REDDIT_BRAND_CONTEXT", "").strip()
REDDIT_CONTENT_GOALS = [
    value.strip()
    for value in os.getenv(
        "REDDIT_CONTENT_GOALS",
        "audience research, subreddit discovery, post strategy, content creation, comment drafting, reporting",
    ).split(",")
    if value.strip()
]
REDDIT_PRIORITY_SUBREDDITS = [
    value.strip()
    for value in os.getenv("REDDIT_PRIORITY_SUBREDDITS", "").split(",")
    if value.strip()
]
REDDIT_TARGET_AUDIENCES = [
    value.strip()
    for value in os.getenv("REDDIT_TARGET_AUDIENCES", "").split(",")
    if value.strip()
]
REDDIT_PROHIBITED_CLAIMS = [
    value.strip()
    for value in os.getenv("REDDIT_PROHIBITED_CLAIMS", "").split(",")
    if value.strip()
]

BOT_USER_ID = ""
BOT_MENTION = ""


def _tool_name(tool: Any) -> str:
    """Return tool name for both object-style and dict-style tool payloads."""
    if isinstance(tool, dict):
        return str(tool.get("name") or "").strip()
    return str(getattr(tool, "name", "") or "").strip()


def _normalized_tool_names(tools: list[Any]) -> list[str]:
    """Collect tool names and filter out missing/empty entries."""
    return [name for name in (_tool_name(tool) for tool in tools) if name]


def get_or_create_session(user_id: str):
    """Create a Composio session for a Slack user or a shared service account."""
    composio_user_id = SHARED_COMPOSIO_USER_ID or user_id
    cache_key = f"{composio_user_id}:{SHARED_CONNECTED_ACCOUNT_ID}"

    if cache_key not in user_sessions:
        create_kwargs: dict[str, Any] = {
            "user_id": composio_user_id,
            "toolkits": DEFAULT_TOOLKITS,
        }
        if SHARED_CONNECTED_ACCOUNT_ID:
            create_kwargs["connected_accounts"] = {
                "reddit": SHARED_CONNECTED_ACCOUNT_ID
            }
        user_sessions[cache_key] = composio.create(**create_kwargs)
    return user_sessions[cache_key]


def get_or_create_memory(user_id: str) -> SQLiteSession:
    """Get or create per-user conversation memory."""
    if user_id not in memory_sessions:
        memory_sessions[user_id] = SQLiteSession(f"slack_{user_id}")
    return memory_sessions[user_id]


def load_access_control() -> dict[str, set[str]]:
    """Load dynamic Slack access control from local storage."""
    if not ACCESS_CONTROL_FILE.exists():
        return {"admins": set(), "allowed_users": set()}

    try:
        payload = json.loads(ACCESS_CONTROL_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {"admins": set(), "allowed_users": set()}

    admins = {
        str(user_id).strip()
        for user_id in payload.get("admins", [])
        if str(user_id).strip()
    }
    allowed_users = {
        str(user_id).strip()
        for user_id in payload.get("allowed_users", [])
        if str(user_id).strip()
    }
    return {"admins": admins, "allowed_users": allowed_users}


def save_access_control(data: dict[str, set[str]]) -> None:
    """Persist dynamic Slack access control to local storage."""
    payload = {
        "admins": sorted(data.get("admins", set())),
        "allowed_users": sorted(data.get("allowed_users", set())),
    }
    ACCESS_CONTROL_FILE.write_text(
        json.dumps(payload, indent=2, ensure_ascii=True) + "\n",
        encoding="utf-8",
    )


def get_allowed_users() -> set[str]:
    """Return the merged static and dynamic allowlist."""
    access_control = load_access_control()
    return set(ENV_ALLOWED_USERS) | set(access_control["allowed_users"])


def get_admin_users() -> set[str]:
    """Return Slack users allowed to manage access from Slack."""
    access_control = load_access_control()
    return set(access_control["admins"])


def validate_shared_account_config() -> None:
    """Fail fast on common shared-account configuration mistakes."""
    if SHARED_CONNECTED_ACCOUNT_ID and not SHARED_COMPOSIO_USER_ID:
        raise RuntimeError(
            "COMPOSIO_SHARED_CONNECTED_ACCOUNT_ID is set, but COMPOSIO_SHARED_USER_ID is missing. "
            "Set both for shared Reddit account mode, or remove COMPOSIO_SHARED_CONNECTED_ACCOUNT_ID."
        )

    if re.fullmatch(r"U[A-Z0-9]{8,}", SHARED_CONNECTED_ACCOUNT_ID):
        raise RuntimeError(
            "COMPOSIO_SHARED_CONNECTED_ACCOUNT_ID looks like a Slack user ID. "
            "Use a Composio connected account ID for Reddit instead, or remove this variable if your shared Composio user has only one Reddit connection."
        )


def _format_prompt_list(values: list[str], fallback: str = "not specified") -> str:
    """Format a list for prompt text."""
    return ", ".join(values) if values else fallback


def _build_reddit_brand_context() -> str:
    """Build environment-driven brand context for the Reddit agent."""
    lines = [
        f"Brand name: {REDDIT_BRAND_NAME}",
        f"Primary goals: {_format_prompt_list(REDDIT_CONTENT_GOALS)}",
        f"Priority subreddits: {_format_prompt_list(REDDIT_PRIORITY_SUBREDDITS, 'discover based on the request')}",
        f"Target audiences: {_format_prompt_list(REDDIT_TARGET_AUDIENCES, 'infer from the request')}",
    ]
    if REDDIT_BRAND_CONTEXT:
        lines.append(f"Brand context: {REDDIT_BRAND_CONTEXT}")
    if REDDIT_PROHIBITED_CLAIMS:
        lines.append(
            f"Prohibited claims or phrasing: {_format_prompt_list(REDDIT_PROHIBITED_CLAIMS)}"
        )
    return "\n".join(lines)


def _workflow_hint(text: str) -> tuple[str, str]:
    """Infer the likely Reddit workflow and response requirements."""
    normalized = text.lower().strip()
    has_reddit_url = bool(re.search(r"https?://(?:www\.)?reddit\.com/\S+", normalized))

    if any(keyword in normalized for keyword in ("excel", "xlsx", "spreadsheet", "export report")):
        return (
            "report_export",
            "Prioritize structured research and include a compact Slack summary plus the required Excel JSON block.",
        )
    if has_reddit_url or "analyze thread" in normalized or "summarize thread" in normalized:
        return (
            "thread_analysis",
            "Analyze the linked Reddit thread, summarize the current context, sentiment, objections, opportunities, and suggested next action.",
        )
    if any(keyword in normalized for keyword in ("where to post", "which subreddit", "best subreddit", "subreddit for")):
        return (
            "subreddit_selection",
            "Recommend the best subreddit options, explain why each fits, note audience fit, posting risk, and which one to prioritize first.",
        )
    if any(keyword in normalized for keyword in ("what to post", "content plan", "content calendar", "campaign plan", "post ideas")):
        return (
            "content_strategy",
            "Create a Reddit content strategy with post angles, recommended subreddits, hooks, timing guidance, and context on why the content should work now.",
        )
    if any(keyword in normalized for keyword in ("draft post", "write post", "create post")):
        return (
            "post_drafting",
            "Draft Reddit-ready post options with title, body, subreddit fit, risk checks, and any comments needed to support the post.",
        )
    if any(keyword in normalized for keyword in ("draft comment", "write comment", "reply comment", "comment draft")):
        return (
            "comment_drafting",
            "Draft concise Reddit comments or replies that match subreddit tone, avoid spam signals, and feel native to the thread.",
        )
    return (
        "research_strategy",
        "Research the topic deeply, explain what is happening now, where it belongs on Reddit, what content should be created, and the best next action.",
    )


def _workflow_output_requirements(workflow: str) -> str:
    """Return response structure guidance for a workflow."""
    workflow_map = {
        "report_export": (
            "Use this structure when possible: short title, *Summary*, *Current Context*, "
            "*Recommended Subreddits*, *Content Opportunities*, *Next Step*."
        ),
        "thread_analysis": (
            "Use this structure when possible: short title, *Summary*, *Current Context*, "
            "*Sentiment*, *Key Pain Points*, *Opportunity*, *Next Step*."
        ),
        "subreddit_selection": (
            "Use this structure when possible: short title, *Summary*, *Best Subreddits*, "
            "*Why These Fit*, *Posting Risks*, *Next Step*."
        ),
        "content_strategy": (
            "Use this structure when possible: short title, *Summary*, *Current Context*, "
            "*Recommended Subreddits*, *Content Plan*, *Risks*, *Next Step*."
        ),
        "post_drafting": (
            "Use this structure when possible: short title, *Goal*, *Recommended Subreddit*, "
            "*Draft Post*, *Support Comment*, *Risks*, *Next Step*."
        ),
        "comment_drafting": (
            "Use this structure when possible: short title, *Summary*, *Thread Context*, "
            "*Draft Comment*, *Why It Works*, *Risks*, *Next Step*."
        ),
    }
    return workflow_map.get(
        workflow,
        "Use this structure when possible: short title, *Summary*, *Current Context*, "
        "*Recommended Subreddits*, *Content Recommendation*, *Risks*, *Next Step*.",
    )


def build_agent(tools: list[Any]) -> Agent:
    """Create the Reddit-specific Slack agent."""
    brand_context = _build_reddit_brand_context()
    return Agent(
        name="Slack Reddit Agent",
        instructions=(
            f"You are the dedicated Reddit Operations Agent for {REDDIT_BRAND_NAME} working inside Slack. "
            "Own the full Reddit workflow: market research, subreddit selection, content strategy, post ideation, "
            "title testing, post drafting, comment drafting, thread analysis, competitor analysis, trend analysis, "
            "and reporting. "
            "Your job is to decide what should be posted, where it should be posted, what the content should say, "
            "what the current Reddit context is, and what action should happen next. "
            "When the user asks a vague question, do not stall. Infer the most useful Reddit workflow, state assumptions briefly, and provide the best operational recommendation. "
            "Check subreddit rules before recommending posts when possible. "
            "If rules cannot be verified, say so clearly and mark the recommendation as lower confidence. "
            "If shared-account mode is enabled, all Reddit actions happen through the company's shared Reddit account, not the Slack user's personal Reddit account. "
            "Never say a Reddit action was completed unless a tool result confirms it. "
            "Ask for explicit confirmation before creating, editing, deleting, submitting, or replying to anything on Reddit. "
            "Before any execution step, show the exact subreddit, title, content, or comment you want to publish. "
            "Avoid spammy, manipulative, fake-organic, brigading, or misleading behavior. Avoid unverified medical or skincare claims. "
            "Keep responses concise, practical, and friendly for Slack. "
            "Format every answer for official Slack Block Kit delivery using Slack mrkdwn, not GitHub Markdown. "
            "When useful, start with one short title line and then leave one blank line before the body so the app can promote that title into a Slack header block. "
            "Use short paragraphs, `*bold*`, `_italic_`, `~strike~`, backticks for inline code, triple backticks for code blocks, `-` for bullets, `1.` for numbered lists, and `<https://example.com|label>` for labeled links. "
            "Prefer Reddit-ops structures such as `*Summary*`, `*Current Context*`, `*Recommended Subreddits*`, `*Content Plan*`, `*Draft Post*`, `*Risks*`, and `*Next Step*` when they fit. "
            "Do not use Markdown headings like `# Heading`, Markdown tables, HTML, nested bullets, or attachment-style formatting. "
            "If the user wants research, explain what conversations are happening now, what the audience cares about, which subreddits matter, and what content angle is most promising. "
            "If the user wants content, provide subreddit fit, hook, title, body, supporting comments, and posting cautions. "
            "If the user wants recommendations on where to post, rank the best subreddit options with reasons and risks. "
            "If the user wants a report, summarize the findings in Slack and prepare the Excel JSON block. "
            "If the user asks for Excel, XLSX, spreadsheet, sheet, table export, or report export, "
            "end your response with one JSON code block and nothing after it. "
            "That JSON must use this exact shape: "
            '{"workbook_name":"reddit_report","sheet_name":"Report","summary":"short summary","columns":["column_a","column_b"],"rows":[{"column_a":"value","column_b":"value"}]}. '
            "Use only flat text, numbers, or booleans in rows. "
            f"Never return more than {MAX_EXPORT_ROWS} rows in that JSON block.\n\n"
            "Brand operating context:\n"
            f"{brand_context}"
        ),
        model=DEFAULT_MODEL,
        tools=tools,
    )


async def run_user_task(user_id: str, task: str) -> str:
    """Run a Slack user task through the Composio-backed agent."""
    session = get_or_create_session(user_id)
    memory = get_or_create_memory(user_id)
    tools = session.tools()
    agent = build_agent(tools)

    result = await asyncio.to_thread(
        Runner.run_sync,
        starting_agent=agent,
        input=task,
        session=memory,
    )
    return str(result.final_output or "No response generated.")


def _clip_text(value: str, limit: int) -> str:
    """Clip prompt text without breaking the request flow."""
    if len(value) <= limit:
        return value
    return value[: limit - 3].rstrip() + "..."


async def _fetch_slack_context(
    client,
    channel: str,
    timestamp: str,
    thread_ts: str | None,
    is_dm: bool,
) -> str:
    """Fetch recent Slack context so the Reddit agent can see surrounding discussion."""
    try:
        if thread_ts:
            response = await client.conversations_replies(
                channel=channel,
                ts=thread_ts,
                limit=SLACK_CONTEXT_MESSAGE_LIMIT,
            )
        else:
            response = await client.conversations_history(
                channel=channel,
                latest=timestamp,
                inclusive=True,
                limit=SLACK_CONTEXT_MESSAGE_LIMIT,
            )
    except Exception:
        return ""

    messages = response.get("messages") or []
    if not messages:
        return ""

    def _ts_value(message: dict[str, Any]) -> float:
        try:
            return float(str(message.get("ts") or "0"))
        except Exception:
            return 0.0

    lines: list[str] = []
    total_length = 0
    for message in sorted(messages, key=_ts_value):
        if message.get("subtype"):
            continue

        raw_text = str(message.get("text") or "").strip()
        if not raw_text:
            continue

        text = raw_text if is_dm else strip_bot_mention(raw_text)
        text = re.sub(r"\s+", " ", text).strip()
        if not text:
            continue

        author = "assistant" if (
            message.get("bot_id") or str(message.get("user") or "").strip() == BOT_USER_ID
        ) else "user"
        line = f"{author}: {text}"
        if total_length + len(line) + 1 > SLACK_CONTEXT_CHAR_LIMIT:
            break
        lines.append(line)
        total_length += len(line) + 1

    return "\n".join(lines)


async def build_reddit_task_prompt(
    client,
    channel: str,
    text: str,
    timestamp: str,
    thread_ts: str | None,
    is_dm: bool,
) -> str:
    """Turn a Slack message into a structured Reddit-ops request for the agent."""
    workflow_name, workflow_hint = _workflow_hint(text)
    context_block = await _fetch_slack_context(
        client=client,
        channel=channel,
        timestamp=timestamp,
        thread_ts=thread_ts,
        is_dm=is_dm,
    )

    prompt_parts = [
        "Handle this as a Reddit operations request from Slack.",
        f"Workflow hint: {workflow_name}",
        f"Workflow goal: {workflow_hint}",
        f"Response format requirement: {_workflow_output_requirements(workflow_name)}",
        (
            "Operational priorities: identify what is happening on Reddit now, decide where content should be posted, "
            "recommend what content should be created, explain why the subreddit fit makes sense, and suggest the next action."
        ),
        f"Slack surface: {'direct message' if is_dm else 'channel mention'}",
        f"Slack context:\n{_clip_text(context_block, SLACK_CONTEXT_CHAR_LIMIT) if context_block else 'No additional Slack context available.'}",
        f"User request:\n{text.strip()}",
    ]
    return "\n\n".join(prompt_parts).strip()


def sanitize_filename(value: str, fallback: str) -> str:
    """Create a filesystem-safe filename stem."""
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip()).strip("._")
    return cleaned or fallback


def sanitize_sheet_name(value: str) -> str:
    """Create an Excel-safe worksheet title."""
    cleaned = re.sub(r"[:\\\\/?*\\[\\]]", " ", value).strip()
    return (cleaned or "Report")[:31]


def _coerce_excel_value(value: Any) -> Any:
    """Normalize nested data for a worksheet cell."""
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=True)


def extract_excel_payload(text: str) -> tuple[str, dict[str, Any] | None]:
    """Extract a JSON export block from an agent response if present."""
    code_block_pattern = re.compile(r"```json\s*(\{.*?\})\s*```", re.DOTALL)

    for match in code_block_pattern.finditer(text):
        candidate = match.group(1).strip()
        try:
            payload = json.loads(candidate)
        except json.JSONDecodeError:
            continue

        if not isinstance(payload, dict):
            continue

        rows = payload.get("rows")
        columns = payload.get("columns")
        if not isinstance(rows, list) or not isinstance(columns, list):
            continue

        message_text = text.replace(match.group(0), "").strip()
        return message_text, payload

    return text, None


def create_excel_report(payload: dict[str, Any]) -> Path:
    """Build an XLSX file from the agent export payload."""
    workbook_name = sanitize_filename(
        str(payload.get("workbook_name") or "reddit_report"),
        fallback="reddit_report",
    )
    sheet_name = sanitize_sheet_name(str(payload.get("sheet_name") or "Report"))
    summary = str(payload.get("summary") or "").strip()
    columns = [str(column).strip() for column in payload.get("columns", []) if str(column).strip()]
    raw_rows = payload.get("rows") or []

    rows: list[dict[str, Any]] = []
    for raw_row in raw_rows[:MAX_EXPORT_ROWS]:
        if isinstance(raw_row, dict):
            rows.append(raw_row)

    if not columns and rows:
        columns = list(rows[0].keys())

    if not columns:
        columns = ["result"]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(columns)

    for header_cell in worksheet[1]:
        header_cell.font = Font(bold=True)

    for row in rows:
        worksheet.append([_coerce_excel_value(row.get(column, "")) for column in columns])

    for index, column in enumerate(columns, start=1):
        cell_values = [str(column)]
        for row in rows:
            cell_values.append(str(_coerce_excel_value(row.get(column, ""))))
        width = min(max(len(value) for value in cell_values) + 2, 60)
        worksheet.column_dimensions[get_column_letter(index)].width = width

    if summary:
        summary_sheet = workbook.create_sheet(title="Summary")
        summary_sheet["A1"] = "Summary"
        summary_sheet["A1"].font = Font(bold=True)
        summary_sheet["A2"] = summary
        summary_sheet.column_dimensions["A"].width = 100

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = EXPORTS_DIR / f"{workbook_name}_{timestamp}.xlsx"
    workbook.save(file_path)
    return file_path


def _normalize_markdown_for_slack(text: str) -> str:
    """Convert common Markdown patterns into Slack-friendly mrkdwn."""
    normalized = text.replace("\r\n", "\n").strip()
    if not normalized:
        return ""

    normalized = re.sub(
        r"\[([^\]]+)\]\((https?://[^\s)]+)\)",
        r"<\2|\1>",
        normalized,
    )
    normalized = re.sub(r"```[A-Za-z0-9_+-]+\n", "```\n", normalized)
    normalized = re.sub(r"(?m)^#{1,6}\s+(.+)$", r"*\1*", normalized)
    normalized = re.sub(r"(?<!\*)\*\*(.+?)\*\*(?!\*)", r"*\1*", normalized)
    normalized = re.sub(r"(?<!_)__(.+?)__(?!_)", r"*\1*", normalized)
    normalized = re.sub(r"~~(.+?)~~", r"~\1~", normalized)
    normalized = re.sub(r"(?m)^\s*[*+]\s+", "- ", normalized)
    normalized = re.sub(r"(?m)^(\s*)(\d+)\)\s+", r"\1\2. ", normalized)
    normalized = re.sub(r"\n{3,}", "\n\n", normalized)
    return normalized.strip()


def _clean_slack_header_text(line: str) -> str:
    """Strip lightweight Slack formatting so a title fits a header block."""
    header = line.strip()
    for pattern, replacement in (
        (r"^\*(.+)\*$", r"\1"),
        (r"^_(.+)_$", r"\1"),
        (r"^~(.+)~$", r"\1"),
        (r"^`(.+)`$", r"\1"),
    ):
        header = re.sub(pattern, replacement, header)
    header = re.sub(r"\s+", " ", header).strip()
    return header[:SLACK_HEADER_TEXT_LIMIT].strip()


def _extract_slack_header_and_body(text: str) -> tuple[str | None, str]:
    """Promote a short first-line title into a Slack header block when possible."""
    lines = text.splitlines()
    if len(lines) < 3:
        return None, text

    title_line = lines[0].strip()
    if not title_line or lines[1].strip():
        return None, text

    if title_line.startswith(("```", "> ", "- ", "* ", "+ ")):
        return None, text

    if re.match(r"^\d+[.)]\s+", title_line):
        return None, text

    header_text = _clean_slack_header_text(title_line)
    if not header_text or len(header_text) > SLACK_HEADER_TEXT_LIMIT:
        return None, text

    body = "\n".join(lines[2:]).strip()
    if not body:
        return None, text

    return header_text, body


def _split_slack_chunks(text: str, limit: int) -> list[str]:
    """Split text into Slack-safe chunks without shredding line structure."""
    if not text.strip():
        return ["No response generated."]

    chunks: list[str] = []
    current_lines: list[str] = []
    current_length = 0

    for line in text.split("\n"):
        line_length = len(line)
        separator = 0 if not current_lines else 1

        if current_lines and current_length + separator + line_length > limit:
            chunks.append("\n".join(current_lines).strip())
            current_lines = []
            current_length = 0

        if line_length <= limit:
            current_lines.append(line)
            current_length += (0 if len(current_lines) == 1 else 1) + line_length
            continue

        if current_lines:
            chunks.append("\n".join(current_lines).strip())
            current_lines = []
            current_length = 0

        start = 0
        while start < line_length:
            chunks.append(line[start : start + limit].strip())
            start += limit

    if current_lines:
        chunks.append("\n".join(current_lines).strip())

    return [chunk for chunk in chunks if chunk] or ["No response generated."]


async def post_chunked_message(
    client,
    channel: str,
    text: str,
    thread_ts: str | None = None,
) -> None:
    """Post Slack-safe chunks to a channel or thread."""
    formatted_text = _normalize_markdown_for_slack(text)
    header_text, body_text = _extract_slack_header_and_body(formatted_text)
    chunks = _split_slack_chunks(body_text or formatted_text, SLACK_BLOCK_TEXT_LIMIT)

    for index, chunk in enumerate(chunks):
        blocks: list[dict[str, Any]] = []
        fallback_parts: list[str] = []

        if index == 0 and header_text:
            blocks.append(
                {
                    "type": "header",
                    "text": {
                        "type": "plain_text",
                        "text": header_text,
                        "emoji": True,
                    },
                }
            )
            fallback_parts.append(header_text)

        blocks.append(
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": chunk,
                    "verbatim": True,
                },
            }
        )
        fallback_parts.append(chunk)

        payload = {
            "channel": channel,
            "text": "\n\n".join(part for part in fallback_parts if part)[:SLACK_MESSAGE_LIMIT],
            "blocks": blocks,
        }
        if thread_ts:
            payload["thread_ts"] = thread_ts
        await client.chat_postMessage(**payload)


async def upload_excel_report(
    client,
    channel: str,
    file_path: Path,
    title: str,
    thread_ts: str | None = None,
) -> None:
    """Upload an Excel file into Slack."""
    payload: dict[str, Any] = {
        "channel": channel,
        "file": str(file_path),
        "filename": file_path.name,
        "title": title,
    }
    if thread_ts:
        payload["thread_ts"] = thread_ts

    await client.files_upload_v2(**payload)


async def add_reaction(client, channel: str, timestamp: str, name: str) -> None:
    """Add a reaction without failing the request flow."""
    try:
        await client.reactions_add(channel=channel, timestamp=timestamp, name=name)
    except Exception:
        pass


async def remove_reaction(client, channel: str, timestamp: str, name: str) -> None:
    """Remove a reaction without failing the request flow."""
    try:
        await client.reactions_remove(channel=channel, timestamp=timestamp, name=name)
    except Exception:
        pass


def strip_bot_mention(text: str) -> str:
    """Remove the bot mention from a Slack message."""
    if not BOT_USER_ID:
        return text.strip()
    pattern = rf"<@{re.escape(BOT_USER_ID)}>"
    return re.sub(pattern, "", text).strip()


def is_allowed_user(user_id: str) -> bool:
    """Allow all users if no allowlist is configured, otherwise restrict access."""
    allowed_users = get_allowed_users()
    admins = get_admin_users()
    if not allowed_users and not admins:
        return False
    return user_id in allowed_users


def is_access_admin(user_id: str) -> bool:
    """Return whether a Slack user can manage the dynamic allowlist."""
    return user_id in get_admin_users()


def _extract_user_ids(text: str) -> list[str]:
    """Extract Slack user IDs from mentions or raw user IDs."""
    mention_ids = re.findall(r"<@([A-Z0-9]+)>", text)
    raw_ids = re.findall(r"\bU[A-Z0-9]{8,}\b", text)
    ordered: list[str] = []
    for user_id in mention_ids + raw_ids:
        if user_id not in ordered:
            ordered.append(user_id)
    return ordered


async def send_access_help(
    client,
    channel: str,
    thread_ts: str | None = None,
) -> None:
    """Explain how to manage the Slack allowlist dynamically."""
    help_text = (
        "*Slack Access Commands*\n"
        "Use these commands in a DM with the bot:\n"
        "- `admin claim` to become the first access admin\n"
        "- `admin list` to see current access admins\n"
        "- `admin add <@user>` to add another admin\n"
        "- `admin remove <@user>` to remove an admin\n"
        "- `allowlist list` to see allowed users\n"
        "- `allowlist add <@user>` to allow a user\n"
        "- `allowlist remove <@user>` to remove a user\n"
        "If `.env` still contains `SLACK_ALLOWED_USERS`, those users remain allowed too."
    )
    await post_chunked_message(client, channel, help_text, thread_ts)


async def handle_access_command(
    client,
    channel: str,
    user_id: str,
    text: str,
    thread_ts: str | None,
    is_dm: bool,
) -> bool:
    """Handle admin and allowlist commands sent from Slack."""
    command_lines = [line.strip() for line in text.splitlines() if line.strip()]
    if len(command_lines) > 1 and all(
        line.lower().startswith(("admin", "allowlist", "access"))
        for line in command_lines
    ):
        for line in command_lines:
            await handle_access_command(
                client,
                channel,
                user_id,
                line,
                thread_ts,
                is_dm,
            )
        return True

    normalized = text.lower().strip()
    reply_thread_ts = None if is_dm else thread_ts
    if not normalized.startswith(("admin", "allowlist", "access")):
        return False

    if not is_dm:
        await post_chunked_message(
            client,
            channel,
            "Manage access in a DM with me so user permissions stay private.",
            reply_thread_ts,
        )
        return True

    if normalized in {"access", "access help", "allowlist help", "admin help"}:
        await send_access_help(client, channel, reply_thread_ts)
        return True

    access_control = load_access_control()
    admins = set(access_control["admins"])
    allowed_users = set(access_control["allowed_users"])
    tokens = text.strip().split()

    if normalized == "admin claim":
        if admins:
            await post_chunked_message(
                client,
                channel,
                "An access admin already exists. Ask a current admin to add you with `admin add <@user>`.",
                reply_thread_ts,
            )
            return True

        admins.add(user_id)
        allowed_users.add(user_id)
        save_access_control({"admins": admins, "allowed_users": allowed_users})
        await post_chunked_message(
            client,
            channel,
            "You are now the first Slack access admin. You have also been added to the allowlist.",
            reply_thread_ts,
        )
        return True

    if normalized == "admin list":
        admin_text = ", ".join(f"<@{admin_id}>" for admin_id in sorted(admins)) or "No Slack admins yet."
        await post_chunked_message(
            client,
            channel,
            f"*Slack Access Admins*\n{admin_text}",
            reply_thread_ts,
        )
        return True

    if normalized == "allowlist list":
        dynamic_text = ", ".join(f"<@{member_id}>" for member_id in sorted(allowed_users))
        env_text = ", ".join(f"`{member_id}`" for member_id in sorted(ENV_ALLOWED_USERS))
        parts = ["*Allowed Slack Users*"]
        parts.append(f"Dynamic: {dynamic_text or 'none'}")
        if ENV_ALLOWED_USERS:
            parts.append(f".env fallback: {env_text}")
        await post_chunked_message(
            client,
            channel,
            "\n".join(parts),
            reply_thread_ts,
        )
        return True

    if user_id not in admins:
        await post_chunked_message(
            client,
            channel,
            "Only Slack access admins can manage allowed users. If this is a new setup, send `admin claim` first.",
            reply_thread_ts,
        )
        return True

    target_ids = _extract_user_ids(text)
    if len(tokens) >= 3 and tokens[0].lower() == "admin" and tokens[1].lower() == "add":
        if not target_ids:
            await post_chunked_message(
                client,
                channel,
                "Use `admin add <@user>`.",
                reply_thread_ts,
            )
            return True

        admins.update(target_ids)
        allowed_users.update(target_ids)
        save_access_control({"admins": admins, "allowed_users": allowed_users})
        await post_chunked_message(
            client,
            channel,
            f"Added admin access for {', '.join(f'<@{target_id}>' for target_id in target_ids)}.",
            reply_thread_ts,
        )
        return True

    if len(tokens) >= 3 and tokens[0].lower() == "admin" and tokens[1].lower() == "remove":
        if not target_ids:
            await post_chunked_message(
                client,
                channel,
                "Use `admin remove <@user>`.",
                reply_thread_ts,
            )
            return True

        remaining_admins = admins - set(target_ids)
        if not remaining_admins:
            await post_chunked_message(
                client,
                channel,
                "You must keep at least one Slack access admin.",
                reply_thread_ts,
            )
            return True

        admins = remaining_admins
        save_access_control({"admins": admins, "allowed_users": allowed_users})
        await post_chunked_message(
            client,
            channel,
            f"Removed admin access for {', '.join(f'<@{target_id}>' for target_id in target_ids)}.",
            reply_thread_ts,
        )
        return True

    if len(tokens) >= 3 and tokens[0].lower() == "allowlist" and tokens[1].lower() == "add":
        if not target_ids:
            await post_chunked_message(
                client,
                channel,
                "Use `allowlist add <@user>`.",
                reply_thread_ts,
            )
            return True

        allowed_users.update(target_ids)
        save_access_control({"admins": admins, "allowed_users": allowed_users})
        await post_chunked_message(
            client,
            channel,
            f"Allowed {', '.join(f'<@{target_id}>' for target_id in target_ids)} to use the bot.",
            reply_thread_ts,
        )
        return True

    if len(tokens) >= 3 and tokens[0].lower() == "allowlist" and tokens[1].lower() == "remove":
        if not target_ids:
            await post_chunked_message(
                client,
                channel,
                "Use `allowlist remove <@user>`.",
                reply_thread_ts,
            )
            return True

        allowed_users.difference_update(target_ids)
        save_access_control({"admins": admins, "allowed_users": allowed_users})
        await post_chunked_message(
            client,
            channel,
            f"Removed {', '.join(f'<@{target_id}>' for target_id in target_ids)} from the allowlist.",
            reply_thread_ts,
        )
        return True

    await send_access_help(client, channel, reply_thread_ts)
    return True


async def send_help(client, channel: str, thread_ts: str | None = None) -> None:
    """Send a simple help message."""
    help_text = (
        "*Reddit Slack Bot*\n"
        "DM me directly for Reddit research, subreddit selection, content planning, drafting, and reporting.\n"
        "Mention me in channels with `@bot_name your task`.\n"
        "Type `tools` to list available Composio Reddit tools.\n"
        "Type `reddit help` to see the best Reddit workflows and prompt formats.\n"
        "Type `access help` in DM to manage allowed Slack users dynamically.\n"
        "Ask for Excel exports with prompts like:\n"
        "- `Analyze r/SkincareAddiction and send the result as an Excel file`\n"
        "- `Export the top 25 Reddit posts about retinol this week into Excel with scores and insights`"
    )
    await post_chunked_message(client, channel, help_text, thread_ts)


async def send_reddit_workflow_help(
    client,
    channel: str,
    thread_ts: str | None = None,
) -> None:
    """Show the highest-value Reddit workflows the bot can handle."""
    help_text = (
        "Reddit Workflows\n\n"
        "*Research*\n"
        "- `research retinol concerns for acne-prone skin`\n"
        "- `summarize what people are discussing in r/SkincareAddiction about sunscreen sticks`\n\n"
        "*Where To Post*\n"
        "- `where should we post a hydration success story on Reddit?`\n"
        "- `which subreddit fits a dermatologist-led skincare AMA?`\n\n"
        "*What To Post*\n"
        "- `what should we post this week for our niacinamide launch?`\n"
        "- `create 5 Reddit post ideas for barrier repair content`\n\n"
        "*Draft Content*\n"
        "- `draft a Reddit post for r/SkincareAddiction about retinol purging`\n"
        "- `draft a helpful comment replying to this concern about tret dryness`\n\n"
        "*Analyze Context*\n"
        "- `analyze this Reddit thread and tell me the sentiment and opportunity: <reddit_url>`\n"
        "- `what is the current context around peptide serums on Reddit right now?`\n\n"
        "*Reports And Excel*\n"
        "- `make a weekly Reddit report for sunscreen discussions and export it to Excel`\n"
        "- `export the top content opportunities for acne education into xlsx`"
    )
    await post_chunked_message(client, channel, help_text, thread_ts)


async def send_tools_list(
    client,
    channel: str,
    user_id: str,
    thread_ts: str | None = None,
) -> None:
    """List the available Composio tools for the current user session."""
    session = get_or_create_session(user_id)
    tool_names = _normalized_tool_names(session.tools() or [])

    if not tool_names:
        await post_chunked_message(
            client,
            channel,
            "No Composio tools are available yet for this user.",
            thread_ts,
        )
        return

    preview = "\n".join(f"• {name}" for name in tool_names[:40])
    if len(tool_names) > 40:
        preview += f"\n…and {len(tool_names) - 40} more"

    await post_chunked_message(
        client,
        channel,
        f"*Available Composio Tools* ({len(tool_names)} total)\n{preview}",
        thread_ts,
    )


def _is_process_running(pid: int) -> bool:
    """Check whether a process ID exists without killing it."""
    try:
        os.kill(pid, 0)
        return True
    except OSError:
        return False


def acquire_single_instance_lock() -> bool:
    """Ensure only one slack_bot.py process is active."""
    if os.path.exists(LOCK_FILE):
        try:
            with open(LOCK_FILE, "r", encoding="utf-8") as file_obj:
                existing_pid = int(file_obj.read().strip())
            if _is_process_running(existing_pid):
                return False
        except Exception:
            pass

    with open(LOCK_FILE, "w", encoding="utf-8") as file_obj:
        file_obj.write(str(os.getpid()))

    def _cleanup_lock() -> None:
        try:
            if os.path.exists(LOCK_FILE):
                with open(LOCK_FILE, "r", encoding="utf-8") as file_obj:
                    pid_in_file = file_obj.read().strip()
                if pid_in_file == str(os.getpid()):
                    os.remove(LOCK_FILE)
        except Exception:
            pass

    atexit.register(_cleanup_lock)
    return True


def build_app() -> AsyncApp:
    """Create and configure the Slack Bolt app."""
    app = AsyncApp(token=os.environ["SLACK_BOT_TOKEN"])

    async def process_event(event: dict[str, Any], logger, is_dm: bool) -> None:
        """Process either a DM message event or a channel app mention event."""
        if event.get("subtype") or event.get("bot_id"):
            return

        user_id = str(event.get("user") or "").strip()
        channel = str(event.get("channel") or "").strip()
        text = str(event.get("text") or "").strip()
        timestamp = str(event.get("ts") or "").strip()

        if not user_id or not channel:
            return

        thread_ts = event.get("thread_ts") or (None if is_dm else timestamp)

        if not is_dm:
            if not BOT_MENTION or BOT_MENTION not in text:
                return
            text = strip_bot_mention(text)

        normalized = text.lower().strip()
        client = app.client

        if await handle_access_command(
            client,
            channel,
            user_id,
            text,
            thread_ts,
            is_dm,
        ):
            return

        if not is_allowed_user(user_id):
            logger.info("Ignoring unauthorized Slack user: %s", user_id)
            await post_chunked_message(
                client,
                channel,
                "You are not allowed to use this bot yet. If this is the first setup, DM me `admin claim`. Otherwise ask an admin to DM me with `allowlist add <@your_name>`.",
                thread_ts,
            )
            return

        if not text:
            await post_chunked_message(
                client,
                channel,
                "Tell me what Reddit task you want help with.",
                thread_ts,
            )
            return

        if normalized in {"help", "/help"}:
            await send_help(client, channel, thread_ts)
            return

        if normalized in {"reddit help", "workflow help", "workflows", "/reddit"}:
            await send_reddit_workflow_help(client, channel, thread_ts)
            return

        if normalized in {"tools", "/tools"}:
            await send_tools_list(client, channel, user_id, thread_ts)
            return

        await add_reaction(client, channel, timestamp, "eyes")

        try:
            task_prompt = await build_reddit_task_prompt(
                client=client,
                channel=channel,
                text=text,
                timestamp=timestamp,
                thread_ts=thread_ts,
                is_dm=is_dm,
            )
            response_text = await run_user_task(user_id, task_prompt)
            message_text, export_payload = extract_excel_payload(response_text)

            if message_text:
                await post_chunked_message(client, channel, message_text, thread_ts)

            if export_payload:
                file_path = create_excel_report(export_payload)
                title = str(export_payload.get("workbook_name") or "reddit_report")
                await upload_excel_report(
                    client,
                    channel,
                    file_path,
                    title=title,
                    thread_ts=thread_ts,
                )

                if not message_text:
                    await post_chunked_message(
                        client,
                        channel,
                        "Excel report created and uploaded.",
                        thread_ts,
                    )

            await remove_reaction(client, channel, timestamp, "eyes")
            await add_reaction(client, channel, timestamp, "white_check_mark")
        except Exception as exc:
            await remove_reaction(client, channel, timestamp, "eyes")
            await add_reaction(client, channel, timestamp, "x")
            await post_chunked_message(
                client,
                channel,
                f"Error: {str(exc)[:1500]}",
                thread_ts,
            )

    @app.event("message")
    async def handle_message_events(body, logger):
        event = body.get("event", {})
        channel_type = str(event.get("channel_type") or "").strip()

        # DMs are handled here. Channel mentions are handled via app_mention
        # to avoid duplicate replies and Bolt warnings.
        if channel_type != "im":
            return

        await process_event(event, logger, is_dm=True)

    @app.event("app_mention")
    async def handle_app_mention_events(body, logger):
        event = body.get("event", {})
        await process_event(event, logger, is_dm=False)

    return app


async def main() -> None:
    """Start the Slack bot using Socket Mode."""
    global BOT_USER_ID, BOT_MENTION

    validate_shared_account_config()

    slack_bot_token = os.getenv("SLACK_BOT_TOKEN")
    slack_app_token = os.getenv("SLACK_APP_TOKEN")

    if not slack_bot_token or not slack_app_token:
        missing = [
            name
            for name, value in {
                "SLACK_BOT_TOKEN": slack_bot_token,
                "SLACK_APP_TOKEN": slack_app_token,
            }.items()
            if not value
        ]
        raise RuntimeError(
            f"Missing required environment variables: {', '.join(missing)}"
        )

    if not acquire_single_instance_lock():
        raise RuntimeError(
            "Another slack_bot.py instance is already running. "
            "Stop the previous process first to avoid duplicate responses."
        )

    app = build_app()

    auth = await app.client.auth_test()
    BOT_USER_ID = str(auth.get("user_id") or "")
    BOT_MENTION = f"<@{BOT_USER_ID}>" if BOT_USER_ID else ""

    print(f"✅ Slack bot connected as {auth.get('user')}")
    print(f"✅ Reddit toolkit default: {', '.join(DEFAULT_TOOLKITS)}")
    if SHARED_COMPOSIO_USER_ID:
        print(f"✅ Shared Reddit account mode: {SHARED_COMPOSIO_USER_ID}")
        if SHARED_CONNECTED_ACCOUNT_ID:
            print(f"✅ Shared connected account: {SHARED_CONNECTED_ACCOUNT_ID}")
    else:
        print("✅ Shared Reddit account mode: disabled")
    print("✅ Slack bot is ready to receive DMs and @mentions")

    handler = AsyncSocketModeHandler(app, slack_app_token)
    await handler.start_async()


if __name__ == "__main__":
    asyncio.run(main())
